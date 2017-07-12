import io

from django.contrib.auth.models import User
from django.shortcuts import reverse
from django.test import TestCase
from openpyxl import load_workbook

from gatheros_event.models import Event, Person
from gatheros_subscription.helpers.subscription import export
from gatheros_subscription.models import Subscription


class BaseExportViewTest(TestCase):
    fixtures = [
        '001_user',
        '003_occupation',
        '005_user',
        '006_person',
        '007_organization',
        '008_member',
        '009_place',
        '010_event',
        '003_form',
        '006_lot',
        '007_subscription',
    ]

    def setUp(self):
        self.event = Event.objects.get(
            slug='seo-e-resultados'
        )

        self.url = reverse(
            'subscription:subscriptions-export',
            kwargs={'event_pk': self.event.pk}
        )


class CanExportViewTest(BaseExportViewTest):
    def setUp(self):
        super(CanExportViewTest, self).setUp()
        self.user = User.objects.get(username="lucianasilva@gmail.com")
        self.client.force_login(self.user)

    def test_get_200(self):
        """
        Retornar a página com o formulário deve ser ok
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<form', 1)

    def test_filter_by_state(self):
        """
        Filtrando por estado
        """
        data = {'format': 'html'}
        response = self.client.get(self.url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<tr', 5)  # Todos os inscritos
        self.assertContains(response, 'Pagina 1 de 2')  # Todos os inscritos

        data = {'format': 'html', 'uf': ['TO']}
        response = self.client.get(self.url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<tr', 1)  # Inscritos de 'TO'

    def test_filter_by_age(self):
        """
        Filtrando por idade
        """
        data = {'format': 'html'}
        response = self.client.get(self.url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<tr', 5)  # Todos os inscritos
        self.assertContains(response, 'Pagina 1 de 2')  # Todos os inscritos

        data = {'format': 'html', 'age_0': 30, 'age_1': 40}
        response = self.client.get(self.url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<tr', 3)  # Inscritos de 30 a 40 anos

        data = {'format': 'html', 'age_0': 40}
        response = self.client.get(self.url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<tr', 1)  # Inscritos maiores de 40 anos

        data = {'format': 'html', 'age_1': 35}
        response = self.client.get(self.url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<tr', 3)  # Inscritos menores de 35 anos

        data = {'format': 'html', 'age_0': '', 'age_1': ''}
        response = self.client.get(self.url, data)
        self.assertEqual(response.status_code, 200)  # Post com campos vazios

    def test_filter_by_gender(self):
        """
        Filtrando por sexo
        """
        data = {'format': 'html'}
        response = self.client.get(self.url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<tr', 5)  # Todos os inscritos
        self.assertContains(response, 'Pagina 1 de 2')  # Todos os inscritos

        data = {'format': 'html', 'gender': [Person.GENDER_MALE]}
        response = self.client.get(self.url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<tr', 5)  # Inscritos Homens

        data = {'format': 'html', 'gender': [Person.GENDER_FEMALE]}
        response = self.client.get(self.url, data)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<tr', 2)  # Inscritos Mulheres

    # noinspection PyMethodMayBeStatic
    def test_pagination(self):
        """
        Testar paginação
        """

    #     pass
    # @Todo: fazer/testar paginação

    def test_download(self):
        """
        Testando o download do arquivo
        """
        data = {'format': 'xls'}
        response = self.client.get(self.url, data)

        # Resposta deve ser OK
        self.assertEqual(response.status_code, 200)

        # Deve estar fazendo o download de um conteúdo
        self.assertEquals(
            response.get('Content-Disposition'),
            "attachment; filename=%s-%s.xls" % (self.event.pk, self.event.slug)
        )

        # Conteúdo deve ser um arquivo xlsx válido
        try:
            output = io.BytesIO(response.content)
            load_workbook(output)

        except Exception as e:
            self.fail(str(e))


class CannotExportViewTest(BaseExportViewTest):
    def setUp(self):
        super(CannotExportViewTest, self).setUp()
        self.user = User.objects.get(username="hugoseabra19@gmail.com")
        self.client.force_login(self.user)

    def test_get_302(self):
        """
        Retornar um erro 302 informando que não pode acessar a exportação
        """
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)


class ExportHelperTest(BaseExportViewTest):
    fixtures = [
        '001_user',
        '003_occupation',
        '005_user',
        '006_person',
        '007_organization',
        '008_member',
        '009_place',
        '010_event',
        '006_lot',
        '007_subscription',
        '008_answer',
        '008_answer',

        '001_default_field',
        '002_default_field_option',
        '003_form',
        '004_field',
        '005_field_option',
        '006_lot',
        '007_subscription',
        '008_answer',
    ]

    def setUp(self):
        event = Event.objects.get(slug='encontro-de-casais-2017')
        self.queryset = Subscription.objects.filter(event=event)

    def test_export(self):
        """
        Garante que o conteúdo retornado pela função de export é um xlsx
        """
        try:
            output = io.BytesIO(export(self.queryset))
            load_workbook(output)

        except Exception as e:
            self.fail(str(e))
