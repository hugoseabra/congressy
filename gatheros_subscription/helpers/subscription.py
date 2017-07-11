import locale
from collections import OrderedDict
from uuid import UUID

from openpyxl import Workbook
from rest_framework.renderers import BaseRenderer
from six import BytesIO, text_type

from gatheros_event.models import Person
from gatheros_subscription.models import Subscription
from gatheros_subscription.serializers import SubscriptionExportSerializer

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')


class ExcelRenderer(BaseRenderer):
    """
    Renderer which serializes to Excel
    """

    media_type = 'application/ms-excel'
    format = 'xlsx'
    level_sep = '.'

    def render(self, data, media_type=None, renderer_context=None):
        """
        Renders serialized *data* into Excel. For a dictionary:
        """
        if data is None:
            return ''

        stream = BytesIO()

        wb = Workbook()
        ws = wb.active

        table = self.tablize(data)

        for row_idx, row in enumerate(table, 1):
            for col_idx, elem in enumerate(row, 1):
                if isinstance(elem, UUID):
                    ws.cell(column=col_idx, row=row_idx, value=str(elem))
                else:
                    ws.cell(column=col_idx, row=row_idx, value=elem)

        wb.save(stream)

        return stream.getvalue()

    def tablize(self, data):
        """
        Convert a list of data into a table.
        """
        if data:

            # First, flatten the data (i.e., convert it to a list of
            # dictionaries that are each exactly one level deep).  The key for
            # each item designates the name of the column that the item will
            # fall into.
            data = self.flatten_data(data)

            # Create a row for each dictionary, filling in columns for which the
            # item has no data with None values.
            rows = []
            cols = data[0].keys()
            for item in data:
                row = []
                for key in cols:
                    row.append(item.get(key, None))
                rows.append(row)

            # Return your "table", with the headers as the first row.
            return [self.get_headers(data)] + rows

        else:
            return []

    def get_headers(self, data):
        """
        Get the set of all unique headers
        :param data: dict
        :return: list
        """
        headers = []
        if data[0]:
            headers = data[0].keys()
        return headers

    def flatten_data(self, data):
        """
        Convert the given data collection to a list of dictionaries that are
        each exactly one level deep. The key for each value in the dictionaries
        designates the name of the column that the value will fall into.
        """
        flat_data = []
        for item in data:
            flat_item = self.flatten_item(item)
            flat_data.append(flat_item)

        return flat_data

    def flatten_item(self, item):
        if isinstance(item, list):
            flat_item = self.flatten_list(item)
        elif isinstance(item, dict):
            flat_item = self.flatten_dict(item)
        else:
            flat_item = {'': item}

        return flat_item

    def nest_flat_item(self, flat_item, prefix):
        """
        Given a "flat item" (a dictionary exactly one level deep), nest all of
        the column headers in a namespace designated by prefix.  For example:
         header... | with prefix... | becomes...
        -----------|----------------|----------------
         'lat'     | 'location'     | 'location.lat'
         ''        | '0'            | '0'
         'votes.1' | 'user'         | 'user.votes.1'
        """
        nested_item = {}
        for header, val in flat_item.items():
            nested_header = self.level_sep.join(
                [prefix, header]) if header else prefix
            nested_item[nested_header] = val
        return nested_item

    def flatten_list(self, l):
        flat_list = {}
        for index, item in enumerate(l):
            index = text_type(index)
            flat_item = self.flatten_item(item)
            nested_item = self.nest_flat_item(flat_item, index)
            flat_list.update(nested_item)
        return flat_list

    def flatten_dict(self, d):
        flat_dict = OrderedDict()
        for key, item in d.items():
            key = str(key)
            flat_item = self.flatten_item(item)
            nested_item = self.nest_flat_item(flat_item, key)
            flat_dict.update(nested_item)
        return flat_dict


class SubscriptionRenderer(ExcelRenderer):
    def get_headers(self, data):
        """
        Get the set of all unique headers

        :param data: dict
        :return: list
        """
        presentation = {
            'count': Subscription._meta.get_field('count').verbose_name,
            'synchronized': 'Sincronizado',
            'origin': Subscription._meta.get_field('origin').verbose_name,
            'lot': 'Lote',
            'code': 'Código de Inscrição',
            'person.uuid': 'Id',
            'person.name': Person._meta.get_field('name').verbose_name,
            'person.gender': Person._meta.get_field('gender').verbose_name,
            'person.city': Person._meta.get_field('city').verbose_name,
        }

        headers = []
        if data[0]:
            for col_name in data[0].keys():
                headers.append(presentation.get(col_name, col_name).title())

        return headers


def export(queryset):
    """
    Exportação do queryset de inscrições em formato xlsx

    :param queryset:
    :return: ByteIO
    """

    # Exportando queryset de inscrições
    queryset = SubscriptionExportSerializer.setup_prefetch(queryset)
    serializer = SubscriptionExportSerializer(data=queryset, many=True)
    serializer.is_valid()

    # Renderizando o resultado em xlsx
    return SubscriptionRenderer().render(serializer.data)
