""" Helper para inscrições. """
import locale
import os

import absoluteuri
from django.conf import settings
from openpyxl import Workbook
from six import BytesIO

from attendance.helpers.attendance import subscription_is_checked
from gatheros_event.helpers.event_business import is_paid_event
from payment.models import Transaction
from survey.models import Answer, Question

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

cached_data = {
    'people': dict(),
    'lots': dict(),
    'subscriptions': dict(),
    'transactions': dict(),
    'addon_products': dict(),
    'addon_services': dict(),
}


def get_object_value(obj, attr, cached_type=None, cached_key=None):
    if not hasattr(obj, attr):
        return ''

    try:
        if cached_type and cached_type in cached_data:
            cached_container = cached_data[cached_type]

            if cached_key and hasattr(obj, cached_key):
                obj_pk = getattr(obj, cached_key)
            else:
                obj_pk = str(obj.pk)

            if obj_pk in cached_container:
                data = cached_container[obj_pk]

                if attr in data:
                    return data[attr]

            else:
                cached_container[obj_pk] = dict()

            value = getattr(obj, attr)
            cached_data[cached_type][obj_pk][attr] = value
            return value

        return getattr(obj, attr)

    except AttributeError:
        return ''


def clean_sheet_title(title):
    """ Limpa título de guia da planilha. """
    title = str(title)

    invalid_chars = (
        '!',
        '@',
        '#',
        '/',
        '&',
        '*',
        ':',
        '?',
    )
    for char in invalid_chars:
        title = title.replace(char, '_')

    return title[0:28]


def export_event_data(event):
    """
    Exportação do queryset de inscrições em formato xlsx

    :param event:
    :return: ByteIO
    """
    stream = BytesIO()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = clean_sheet_title('Participantes')

    subscriptions = event.subscriptions.filter(
        completed=True,
        test_subscription=False,
    )

    _export_subscriptions(ws1, subscriptions)

    if is_paid_event(event):
        _export_payments(wb.create_sheet(title='Pagamentos'), event)

    for ev_survey in event.surveys.all():
        title = clean_sheet_title(
            'Formulário-{}'.format(ev_survey.survey.name)
        )
        _export_survey_answers(wb.create_sheet(title=title),
                               ev_survey,
                               subscriptions)

    for lot_category in event.lot_categories.all():
        cat_name = lot_category.name

        products_queryset = lot_category.product_optionals
        services_queryset = lot_category.service_optionals

        if products_queryset.count():
            num_subs = products_queryset.filter(
                subscription_products__subscription__completed=True,
                subscription_products__subscription__test_subscription=False
            ).count()

            if num_subs > 0:
                sheet_name = 'Opcionais - {}'.format(cat_name)
                title = clean_sheet_title(sheet_name)
                worksheet = wb.create_sheet(title=title)

                _export_addon_products(worksheet, products_queryset.all())

        if services_queryset.count():
            num_subs = services_queryset.filter(
                subscription_services__subscription__completed=True,
                subscription_services__subscription__test_subscription=False
            ).count()

            if num_subs > 0:
                sheet_name = 'Ativ. Extras - {}'.format(cat_name)
                title = clean_sheet_title(sheet_name)
                worksheet = wb.create_sheet(title=title)

                _export_addon_services(worksheet, services_queryset.all())

    wb.save(stream)

    return stream.getvalue()


def _export_subscriptions(worksheet, subscriptions):
    """ Exporta dados de inscrição. """

    worksheet.append([
        'NÚMERO DE INSCRIÇÃO',
        'CÓDIGO DA INSCRIÇÃO',
        'CATEGORIA DE PARTICIPANTE',
        'LOTE',
        'STATUS',
        'ATENDIDO',
        'SEXO',
        'NOME',
        'TIPO DE DOCUMENTO',
        'NÚMERO DO DOCUMENTO',
        'DATA NASC',
        'IDADE',
        'EMAIL',
        'PHONE',
        'RUA/LOGRADOURO',
        'COMPLEMENTO',
        'NUMERO',
        'BAIRRO',
        'CEP',
        'CIDADE',
        'UF',
        'PAÍS',
        'INSTITUICAO/EMPRESA',
        'CNPJ',
        'FUNÇÃO/CARGO',
        'TAG DE AGRUPAMENTO',
        'INFO. PARA CRACHÁ',
        'OBS',
        'CRIADO EM',
    ])

    collector = {}
    row_idx = 1
    for sub in subscriptions:
        if row_idx not in collector:
            collector[row_idx] = []

        person = get_object_value(
            obj=sub,
            attr='person',
            cached_type='subscriptions',
        )

        lot = get_object_value(
            obj=sub,
            attr='lot',
            cached_type='subscriptions',
        )

        lot_category = get_object_value(
            obj=lot,
            attr='category',
            cached_type='lots',
        )

        city = get_object_value(
            obj=person,
            attr='city',
            cached_type='people',
        )

        collector[row_idx].append(get_object_value(
            obj=sub,
            attr='event_count',
            cached_type='subscriptions',
        ))

        collector[row_idx].append(get_object_value(
            obj=sub,
            attr='code',
            cached_type='subscriptions',
        ))

        collector[row_idx].append(lot_category.name.upper())

        collector[row_idx].append(lot.name.upper())

        collector[row_idx].append(sub.get_status_display())

        if subscription_is_checked(sub.pk):
            is_checked = "Sim"
        else:
            is_checked = "Não"

        collector[row_idx].append(is_checked)
        collector[row_idx].append(person.get_gender_display())

        collector[row_idx].append(get_object_value(person, 'name'))

        country = get_object_value(person, 'country')
        if country == 'BR':
            doc_type = 'CPF'
            doc_num = get_object_value(person, 'cpf')
        else:
            doc_type = get_object_value(person, 'international_doc_type')
            doc_num = get_object_value(person, 'international_doc')

        collector[row_idx].append(doc_type)
        collector[row_idx].append(doc_num)

        if person.birth_date:
            collector[row_idx].append(person.birth_date.strftime('%d/%m/%Y'))
            collector[row_idx].append(person.age)
        else:
            collector[row_idx].append('')
            collector[row_idx].append('')

        collector[row_idx].append(get_object_value(person, 'email'))
        collector[row_idx].append(get_object_value(person, 'phone'))
        if country == 'BR':
            street = get_object_value(person, 'street')
            number = get_object_value(person, 'number')
            village = get_object_value(person, 'village')
            zip_code = get_object_value(person, 'zip_code')
            city_name = get_object_value(city, 'name')
            uf = get_object_value(city, 'uf')

        else:
            street = get_object_value(person, 'address_international')
            number = ''
            village = ''
            zip_code = get_object_value(person, 'zip_code_international')
            city_name = get_object_value(person, 'city_international')
            uf = get_object_value(person, 'state_international')

        collector[row_idx].append(street)
        collector[row_idx].append(get_object_value(person, 'complement'))
        collector[row_idx].append(number)
        collector[row_idx].append(village)
        collector[row_idx].append(zip_code)
        collector[row_idx].append(city_name)
        collector[row_idx].append(uf)
        collector[row_idx].append(country)
        collector[row_idx].append(get_object_value(person, 'institution'))
        collector[row_idx].append(get_object_value(person, 'institution_cnpj'))
        collector[row_idx].append(get_object_value(person, 'function'))
        collector[row_idx].append(get_object_value(sub, 'tag_group'))
        collector[row_idx].append(get_object_value(sub, 'tag_info'))
        collector[row_idx].append(get_object_value(sub, 'obs'))
        collector[row_idx].append(sub.created.strftime('%d/%m/%Y %H:%M:%S'))

        row_idx += 1

    for row in collector.keys():
        worksheet.append(collector[row])


def _export_payments(worksheet, event):
    worksheet.append([
        'NÚMERO DE INSCRIÇÃO',
        'CÓDIGO DA INSCRIÇÃO',
        'NOME',
        'TIPO',
        'STATUS',
        'PAGARME ID',
        'DATA PAGAMENTO',
        'VALOR INSCRICAO (R$)',
        'VALOR A RECEBER (R$)',
    ])

    transactions = Transaction.objects.filter(subscription__event_id=event.pk)

    collector = {}
    row_idx = 1
    for transaction in transactions:
        if row_idx not in collector:
            collector[row_idx] = []

        sub = get_object_value(
            obj=transaction,
            attr='subscription',
            cached_type='transactions',
        )

        person = get_object_value(
            obj=sub,
            attr='person',
            cached_type='subscriptions',
        )

        created = transaction.date_created.strftime('%d/%m/%Y %H:%M:%S')

        collector[row_idx].append(get_object_value(sub, 'event_count'))
        collector[row_idx].append(get_object_value(sub, 'code'))

        collector[row_idx].append(person.name.upper())

        collector[row_idx].append(transaction.get_type_display())
        collector[row_idx].append(transaction.get_status_display())
        collector[row_idx].append(transaction.pagarme_id)
        collector[row_idx].append(created)
        collector[row_idx].append(transaction.lot.get_calculated_price())
        collector[row_idx].append(transaction.transaction_liquid_amount)

        row_idx += 1

    for row in collector.keys():
        worksheet.append(collector[row])


def _export_survey_answers(worksheet, event_survey, subscriptions):
    """
    Exporta as respostas de survey.
    """

    columns = [
        'NÚMERO DE INSCRIÇÃO',
        'CÓDIGO DA INSCRIÇÃO',
        'NOME',
        'E-MAIL',
    ]

    survey = event_survey.survey
    event = event_survey.event

    questions = survey.questions.all().order_by('order')

    # Lista a ser consultada para pegar a sequência de colunas de perguntas.
    question_pks = []
    for question in questions:
        # Adiciona coluna da pergunta
        column_name = str(question.label).upper()

        if question.required is True:
            column_name = '* {}'.format(column_name)

        if question.active is False:
            column_name += ' (desativado)'

        columns.append(column_name)

        # Marca coluna da pergunta pelo índice da lista
        question_pks.append(question.pk)

    worksheet.append(columns)

    collector = {}
    row_idx = 1
    for sub in subscriptions:

        person = get_object_value(
            obj=sub,
            attr='person',
            cached_type='subscriptions',
        )

        if row_idx not in collector:
            collector[row_idx] = []

        author = None
        if person.user_id:
            authors = person.user.authors.filter(
                survey_id=survey.pk
            )
            if authors.count():
                author = authors.last()
        elif sub.author_id:
            author = sub.author

        if not author:
            continue

        answers = Answer.objects.filter(
            question__survey_id=survey.pk,
            author_id=author.pk,
        ).order_by('question__order')

        if not answers:
            continue

        upload_file_types = dict()
        upload_file_types[Question.FIELD_INPUT_FILE_PDF] = 'pdfs'
        upload_file_types[Question.FIELD_INPUT_FILE_IMAGE] = 'images'

        answers_values = {}
        for answer in answers:
            question = answer.question
            answer_value = answer.human_display
            if not answer_value:
                answers_values[question.pk] = answer_value
                continue

            if question.type in upload_file_types:
                file_url = os.path.join(
                    settings.MEDIA_URL,
                    'survey',
                    upload_file_types[question.type],
                    str(question.pk),
                    answer_value,
                )
                file_url = absoluteuri.build_absolute_uri(file_url)
                answers_values[question.pk] = \
                    '=HYPERLINK("{}")'.format(file_url)
            else:
                answers_values[question.pk] = answer_value

        collector[row_idx].append(get_object_value(sub, 'event_count'))
        collector[row_idx].append(get_object_value(sub, 'code'))
        collector[row_idx].append(person.name.upper())
        collector[row_idx].append(get_object_value(sub, 'email'))

        # Varrer por pergunta para depois encontrar as respostas.
        for question in questions:
            # Se não há resposta, deixar em branco.
            collector[row_idx].append(answers_values.get(question.pk, '-'))

        row_idx += 1

    for row in collector.keys():
        worksheet.append(collector[row])


def _export_addon_products(worksheet, products):
    """
    Exporta Insrições de Opcionais de um evento.
    """
    columns = [
        'NÚMERO DE INSCRIÇÃO',
        'CÓDIGO DA INSCRIÇÃO',
        'NOME',
        'E-MAIL',
        'CATEGORIA',
        'NOME DO OPCIONAL',
        'STATUS',
        'DATA PAGAMENTO',
        'VALOR INSCRICAO (R$)',
        'VALOR A RECEBER (R$)',
    ]

    collector = {}
    row_idx = 1

    for product in products:
        subs = product.subscription_products.filter(
            subscription__completed=True,
            subscription__test_subscription=False,
        )

        for addon_sub in subs:
            sub = get_object_value(
                obj=addon_sub,
                attr='subscription',
                cached_type='addon_products',
            )

            person = get_object_value(
                obj=sub,
                attr='person',
                cached_type='subscriptions',
            )

            lot = get_object_value(
                obj=sub,
                attr='lot',
                cached_type='subscriptions',
            )

            lot_category = get_object_value(
                obj=lot,
                attr='category',
                cached_type='lots',
            )

            optional = get_object_value(
                obj=addon_sub,
                attr='optional',
                cached_type='addon_products',
            )

            created = addon_sub.created.strftime('%d/%m/%Y %H:%M:%S')

            if row_idx not in collector:
                collector[row_idx] = []

            collector[row_idx].append(get_object_value(sub, 'event_count'))
            collector[row_idx].append(get_object_value(sub, 'code'))
            collector[row_idx].append(person.name.upper())
            collector[row_idx].append(get_object_value(sub, 'email'))
            collector[row_idx].append(lot_category.name)
            collector[row_idx].append(optional.name)
            collector[row_idx].append(sub.get_status_display())
            collector[row_idx].append(created)
            collector[row_idx].append(optional.price)
            collector[row_idx].append(optional.liquid_price)

            row_idx += 1

    rows = collector.keys()

    if rows:
        worksheet.append(columns)
        [worksheet.append(collector[row]) for row in rows]


def _export_addon_services(worksheet, services):
    """
    Exporta Insrições de atividades extras.
    """
    columns = [
        'NÚMERO DE INSCRIÇÃO',
        'CÓDIGO DA INSCRIÇÃO',
        'NOME',
        'E-MAIL',
        'CATEGORIA',
        'TEMA',
        'NOME DA ATIVIDADE EXTRA',
        'STATUS',
        'DATA PAGAMENTO',
        'VALOR INSCRICAO (R$)',
        'VALOR A RECEBER (R$)',
    ]

    collector = {}
    row_idx = 1

    for service in services:
        subs = service.subscription_services.filter(
            subscription__completed=True,
            subscription__test_subscription=False,
        )

        for addon_sub in subs:

            sub = get_object_value(
                obj=addon_sub,
                attr='subscription',
                cached_type='addon_services',
            )

            person = get_object_value(
                obj=sub,
                attr='person',
                cached_type='subscriptions',
            )

            lot = get_object_value(
                obj=sub,
                attr='lot',
                cached_type='subscriptions',
            )

            lot_category = get_object_value(
                obj=lot,
                attr='category',
                cached_type='lots',
            )

            optional = get_object_value(
                obj=addon_sub,
                attr='optional',
                cached_type='addon_services',
            )

            theme = get_object_value(
                obj=optional,
                attr='theme',
                cached_type='addon_services',
            )

            created = addon_sub.created.strftime('%d/%m/%Y %H:%M:%S')

            if row_idx not in collector:
                collector[row_idx] = []

            collector[row_idx].append(get_object_value(sub, 'event_count'))
            collector[row_idx].append(get_object_value(sub, 'code'))
            collector[row_idx].append(person.name.upper())
            collector[row_idx].append(person.email)
            collector[row_idx].append(lot_category.name)
            collector[row_idx].append(theme.name)
            collector[row_idx].append(optional.name)
            collector[row_idx].append(sub.get_status_display())
            collector[row_idx].append(created)
            collector[row_idx].append(optional.price)
            collector[row_idx].append(optional.liquid_price)

            row_idx += 1

    rows = collector.keys()

    if rows:
        worksheet.append(columns)
        [worksheet.append(collector[row]) for row in rows]
