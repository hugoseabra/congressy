import locale

from django.db.models import QuerySet
from openpyxl import Workbook
from six import BytesIO

from attendance.models import Checkin, Checkout
from gatheros_subscription.helpers.export import clean_sheet_title, \
    get_object_value
from gatheros_subscription.models import Subscription

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')


def export_attendance(service):
    """
    Exportação de inscrições de um serviço de atendimento em formato xlsx
    :return: ByteIO
    """
    stream = BytesIO()

    wb = Workbook()

    # ===================== PRESENTES =========================================

    ws1 = wb.active
    ws1.title = clean_sheet_title('Presentes')

    checkins = Checkin.objects.filter(
        attendance_service_id=service.pk,
        checkout__isnull=True,
    )

    subscriptions = list()

    for checkin in checkins:
        if checkin.subscription not in subscriptions:
            subscriptions.append(checkin.subscription)

    subscriptions = sort_subscriptions(subscriptions)

    if len(subscriptions) == 0:
        ws1.append(["Nenhuma inscrição com presença"])
    else:
        export_attendance_checkins(
            worksheet=ws1,
            subscriptions=subscriptions,
            attendance=service
        )

    # ===================== AUSENTES ==========================================

    ws2 = wb.create_sheet(title='Ausentes (checkouts realizados)')

    checkouts = Checkin.objects.filter(
        attendance_service=service,
        checkout__isnull=False,
    ).exclude(
        subscription_id__in=[str(s.pk) for s in subscriptions]
    ).order_by('registration', 'created_on')

    subscriptions = list()

    for checkin in checkouts:
        if checkin.subscription not in subscriptions:
            subscriptions.append(checkin.subscription)

    subscriptions = sort_subscriptions(subscriptions)

    if len(subscriptions) == 0:
        ws2.append(["Nenhuma inscrição com ausencia"])
    else:
        export_attendance_checkouts(
            worksheet=ws2,
            subscriptions=subscriptions,
            attendance=service,
        )

    # ===================== AUDITORIA =========================================

    ws3 = wb.create_sheet(title='Auditoria')

    checkins = Checkin.objects.filter(
        attendance_service=service,
    ).order_by('registration', 'created_on')

    subscriptions = list()

    for checkin in checkins:
        if checkin.subscription not in subscriptions:
            subscriptions.append(checkin.subscription)

    subscriptions = sort_subscriptions(subscriptions)

    if len(subscriptions) == 0:
        ws3.append(["Nenhuma inscrição neste serviço de atendimento"])
    else:
        export_attendance_auditing(
            worksheet=ws3,
            subscriptions=subscriptions,
            attendance=service,
        )

    wb.save(stream)
    return stream.getvalue()


def export_attendance_checkins(worksheet, subscriptions, attendance):
    """ Exporta dados de inscrição. """

    worksheet.append([
        'NÚMERO DE INSCRIÇÃO',
        'CÓDIGO DA INSCRIÇÃO',
        'CATEGORIA DE PARTICIPANTE',
        'LOTE',
        'STATUS',
        'SEXO',
        'NOME',
        'TIPO DE DOCUMENTO',
        'NÚMERO DO DOCUMENTO',
        'DATA NASC',
        'IDADE',
        'EMAIL',
        'PHONE',
        'RUA/LOGRADOURO',
        'COMPLEMENTO',
        'NUMERO',
        'BAIRRO',
        'CEP',
        'CIDADE',
        'UF',
        'PAÍS',
        'INSTITUICAO/EMPRESA',
        'CNPJ',
        'FUNÇÃO/CARGO',
        'TAG DE AGRUPAMENTO',
        'INFO. PARA CRACHÁ',
        'OBS',
        'CRIADO EM',
        'HORA REGISTRADA',
    ])

    collector = {}
    row_idx = 1
    for sub in subscriptions:

        checkins = Checkin.objects.filter(
            subscription=sub,
            checkout__isnull=True,
            attendance_service=attendance,
        )

        if row_idx not in collector:
            collector[row_idx] = []

        person = get_object_value(
            obj=sub,
            attr='person',
            cached_type='subscriptions',
        )

        lot = get_object_value(
            obj=sub,
            attr='lot',
            cached_type='subscriptions',
        )

        lot_category = get_object_value(
            obj=lot,
            attr='category',
            cached_type='lots',
        )

        city = get_object_value(
            obj=person,
            attr='city',
            cached_type='people',
        )

        collector[row_idx].append(get_object_value(
            obj=sub,
            attr='event_count',
            cached_type='subscriptions',
        ))

        collector[row_idx].append(get_object_value(
            obj=sub,
            attr='code',
            cached_type='subscriptions',
        ))

        collector[row_idx].append(lot_category.name)

        collector[row_idx].append(lot.name)

        collector[row_idx].append(sub.get_status_display())
        collector[row_idx].append(person.get_gender_display())

        collector[row_idx].append(get_object_value(person, 'name'))

        country = get_object_value(person, 'country')
        if country == 'BR':
            doc_type = 'CPF'
            doc_num = get_object_value(person, 'cpf')
        else:
            doc_type = get_object_value(person, 'international_doc_type')
            doc_num = get_object_value(person, 'international_doc')

        collector[row_idx].append(doc_type)
        collector[row_idx].append(doc_num)

        if person.birth_date:
            collector[row_idx].append(person.birth_date.strftime('%d/%m/%Y'))
            collector[row_idx].append(person.age)
        else:
            collector[row_idx].append('')
            collector[row_idx].append('')

        collector[row_idx].append(get_object_value(person, 'email'))
        collector[row_idx].append(get_object_value(person, 'phone'))
        if country == 'BR':
            street = get_object_value(person, 'street')
            number = get_object_value(person, 'number')
            village = get_object_value(person, 'village')
            zip_code = get_object_value(person, 'zip_code')
            city_name = get_object_value(city, 'name')
            uf = get_object_value(city, 'uf')

        else:
            street = get_object_value(person, 'address_international')
            number = ''
            village = ''
            zip_code = get_object_value(person, 'zip_code_international')
            city_name = get_object_value(person, 'city_international')
            uf = get_object_value(person, 'state_international')

        collector[row_idx].append(street)
        collector[row_idx].append(get_object_value(person, 'complement'))
        collector[row_idx].append(number)
        collector[row_idx].append(village)
        collector[row_idx].append(zip_code)
        collector[row_idx].append(city_name)
        collector[row_idx].append(uf)
        collector[row_idx].append(country)
        collector[row_idx].append(get_object_value(person, 'institution'))
        collector[row_idx].append(get_object_value(person, 'institution_cnpj'))
        collector[row_idx].append(get_object_value(person, 'function'))
        collector[row_idx].append(get_object_value(sub, 'tag_group'))
        collector[row_idx].append(get_object_value(sub, 'tag_info'))
        collector[row_idx].append(get_object_value(sub, 'obs'))
        collector[row_idx].append(sub.created.strftime('%d/%m/%Y %H:%M:%S'))

        if checkins.count() > 1:

            checkin_str = ''

            for i in range(1, len(checkins)):

                instance = checkins[i - 1]

                if instance.registration:
                    t = instance.registration.strftime('%d/%m/%Y %H:%M:%S')
                else:
                    t = instance.created_on.strftime('%d/%m/%Y %H:%M:%S')

                checkin_str += "{}º ás {} \n".format(i, t)

            # Removing last '\n'
            checkin_str = checkin_str[:-1]

        elif checkins.count() == 1:

            instance = checkins[0]

            if instance.registration:
                t = instance.registration.strftime('%d/%m/%Y %H:%M:%S')
            else:
                t = instance.created_on.strftime('%d/%m/%Y %H:%M:%S')

            checkin_str = "{}".format(t)
        else:
            raise Exception("Trying to export checkedin subscription without "
                            "checkin: {}".format(str(sub.pk)))

        collector[row_idx].append(checkin_str)

        row_idx += 1

    for row in collector.keys():
        worksheet.append(collector[row])


def export_attendance_checkouts(worksheet, subscriptions, attendance):
    """ Exporta dados de inscrição. """

    worksheet.append([
        'NÚMERO DE INSCRIÇÃO',
        'CÓDIGO DA INSCRIÇÃO',
        'CATEGORIA DE PARTICIPANTE',
        'LOTE',
        'STATUS',
        'SEXO',
        'NOME',
        'TIPO DE DOCUMENTO',
        'NÚMERO DO DOCUMENTO',
        'DATA NASC',
        'IDADE',
        'EMAIL',
        'PHONE',
        'RUA/LOGRADOURO',
        'COMPLEMENTO',
        'NUMERO',
        'BAIRRO',
        'CEP',
        'CIDADE',
        'UF',
        'PAÍS',
        'INSTITUICAO/EMPRESA',
        'CNPJ',
        'FUNÇÃO/CARGO',
        'TAG DE AGRUPAMENTO',
        'INFO. PARA CRACHÁ',
        'OBS',
        'CRIADO EM',
        'HORA REGISTRADA',
    ])

    collector = {}
    row_idx = 1
    for sub in subscriptions:

        checkout = Checkout.objects.filter(
            checkin__subscription_id=sub.pk,
            checkin__attendance_service_id=attendance.pk,
        ).order_by('registration', 'created_by').last()

        if row_idx not in collector:
            collector[row_idx] = []

        person = get_object_value(
            obj=sub,
            attr='person',
            cached_type='subscriptions',
        )

        lot = get_object_value(
            obj=sub,
            attr='lot',
            cached_type='subscriptions',
        )

        lot_category = get_object_value(
            obj=lot,
            attr='category',
            cached_type='lots',
        )

        city = get_object_value(
            obj=person,
            attr='city',
            cached_type='people',
        )

        collector[row_idx].append(get_object_value(
            obj=sub,
            attr='event_count',
            cached_type='subscriptions',
        ))

        collector[row_idx].append(get_object_value(
            obj=sub,
            attr='code',
            cached_type='subscriptions',
        ))

        collector[row_idx].append(lot_category.name)

        collector[row_idx].append(lot.name)

        collector[row_idx].append(sub.get_status_display())
        collector[row_idx].append(person.get_gender_display())

        collector[row_idx].append(get_object_value(person, 'name'))

        country = get_object_value(person, 'country')
        if country == 'BR':
            doc_type = 'CPF'
            doc_num = get_object_value(person, 'cpf')
        else:
            doc_type = get_object_value(person, 'international_doc_type')
            doc_num = get_object_value(person, 'international_doc')

        collector[row_idx].append(doc_type)
        collector[row_idx].append(doc_num)

        if person.birth_date:
            collector[row_idx].append(person.birth_date.strftime('%d/%m/%Y'))
            collector[row_idx].append(person.age)
        else:
            collector[row_idx].append('')
            collector[row_idx].append('')

        collector[row_idx].append(get_object_value(person, 'email'))
        collector[row_idx].append(get_object_value(person, 'phone'))
        if country == 'BR':
            street = get_object_value(person, 'street')
            number = get_object_value(person, 'number')
            village = get_object_value(person, 'village')
            zip_code = get_object_value(person, 'zip_code')
            city_name = get_object_value(city, 'name')
            uf = get_object_value(city, 'uf')

        else:
            street = get_object_value(person, 'address_international')
            number = ''
            village = ''
            zip_code = get_object_value(person, 'zip_code_international')
            city_name = get_object_value(person, 'city_international')
            uf = get_object_value(person, 'state_international')

        collector[row_idx].append(street)
        collector[row_idx].append(get_object_value(person, 'complement'))
        collector[row_idx].append(number)
        collector[row_idx].append(village)
        collector[row_idx].append(zip_code)
        collector[row_idx].append(city_name)
        collector[row_idx].append(uf)
        collector[row_idx].append(country)
        collector[row_idx].append(get_object_value(person, 'institution'))
        collector[row_idx].append(get_object_value(person, 'institution_cnpj'))
        collector[row_idx].append(get_object_value(person, 'function'))
        collector[row_idx].append(get_object_value(sub, 'tag_group'))
        collector[row_idx].append(get_object_value(sub, 'tag_info'))
        collector[row_idx].append(get_object_value(sub, 'obs'))
        collector[row_idx].append(sub.created.strftime('%d/%m/%Y %H:%M:%S'))

        instance = checkout

        if instance.registration:
            t = instance.registration.strftime('%d/%m/%Y %H:%M:%S')
        else:
            t = instance.created_on.strftime('%d/%m/%Y %H:%M:%S')

        collector[row_idx].append("{}".format(t))

        row_idx += 1

    for row in collector.keys():
        worksheet.append(collector[row])


def export_attendance_auditing(worksheet, subscriptions, attendance):
    """ Exporta dados de inscrição. """

    worksheet.append([
        'CÓDIGO DA INSCRIÇÃO',
        'NOME',
        'CHECKIN',
        'CHECKOUT',
    ])

    collector = {}
    row_idx = 1

    for sub in subscriptions:

        checkins = Checkin.objects.filter(
            subscription_id=sub.pk,
            attendance_service_id=attendance.id,
        ).order_by('registration', 'created_by')

        for checkin in checkins:

            if row_idx not in collector:
                collector[row_idx] = []

            person = get_object_value(
                obj=sub,
                attr='person',
                cached_type='subscriptions',
            )

            collector[row_idx].append(get_object_value(
                obj=sub,
                attr='code',
                cached_type='subscriptions',
            ))

            collector[row_idx].append(get_object_value(person, 'name'))

            instance = checkin

            if instance.registration:
                t = instance.registration.strftime('%d/%m/%Y %H:%M:%S')
            else:
                t = instance.created_on.strftime('%d/%m/%Y %H:%M:%S')

            collector[row_idx].append("{}".format(t))

            try:
                checkout = Checkout.objects.get(checkin_id=checkin.pk)

                if checkout.registration:
                    t = checkout.registration.strftime('%d/%m/%Y %H:%M:%S')
                else:
                    t = checkout.created_on.strftime('%d/%m/%Y %H:%M:%S')

                collector[row_idx].append('{}'.format(t))
            except Checkout.DoesNotExist:
                collector[row_idx].append('---')

            row_idx += 1

    for row in collector.keys():
        worksheet.append(collector[row])


def sort_subscriptions(subscriptions: list) -> QuerySet:
    return Subscription.objects.filter(
        uuid__in=[str(sub.pk) for sub in subscriptions]).order_by(
        'person__name')
