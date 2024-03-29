import abc
import json
import locale
from collections import OrderedDict, namedtuple
from csv import DictReader

from openpyxl import Workbook, styles
from openpyxl.comments import Comment
from openpyxl.styles import colors
from six import BytesIO

from core.util import merge_lists_ignore_duplicates
from gatheros_subscription.models import Lot, FormConfig
from importer.constants import KEY_MAP
from importer.helpers import (
    get_mapping_from_csv_key,
    MappingNotFoundError,
    get_keys_mapping_dict,
)
from survey.models import Question


class XLSPersister(abc.ABC):

    @abc.abstractmethod
    def _get_header(self) -> list:
        pass

    @abc.abstractmethod
    def make(self) -> bytes:
        pass

    @staticmethod
    def colnum_string(n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string


class CSVMixin(object):

    def __init__(self, csv_file_path) -> None:
        self.file_path = csv_file_path
        self._reader = None

    def _get_reader(self) -> DictReader:
        self._reader = DictReader(open(self.file_path, 'r'))

        return self._reader


class XLSErrorPersister(CSVMixin, XLSPersister):

    def __init__(self, csv_file_path) -> None:
        super().__init__(csv_file_path)
        self.redFill = styles.PatternFill(
            patternType='solid',
            fill_type='solid',
            start_color=colors.RED,
            end_color=colors.RED,
        )

    def _get_header(self, survey=None) -> list:

        form_keys = []
        headers = []
        survey_headers = []
        if survey:
            questions = Question.objects.filter(
                survey=survey,
            ).order_by('order')

            for question in questions:
                cleaned_label = question.label.lower().strip()
                survey_headers.append(cleaned_label)

        for line in self._get_reader():

            raw_data = json.loads(line['raw_data'])
            errors = json.loads(line['errors'])

            raw_data_keys = raw_data.keys()
            for key in raw_data_keys:
                if key not in form_keys:
                    form_keys.append(key)

            error_keys = errors.keys()
            for key in error_keys:
                if key not in form_keys:
                    form_keys.append(key)

        for key in form_keys:
            if key not in survey_headers:
                headers.append(KEY_MAP[key]['csv_keys'][0])

        return merge_lists_ignore_duplicates(headers, survey_headers)

    def make(self, survey=None) -> bytes:

        raw_headers = self._get_header(survey)
        headers = list()
        for header in raw_headers:
            headers.append(header.upper())

        cell_mapping = OrderedDict()
        i = 1
        for key in headers:
            letter = self.colnum_string(i)
            cell_mapping.update({key: letter.upper()})
            i += 1

        stream = BytesIO()

        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

        wb = Workbook()

        ws1 = wb.active
        ws1.title = 'Inscrições com erros'

        ws1.append(headers)

        line_counter = 2

        for line in self._get_reader():
            raw_data = json.loads(line['raw_data'])
            errors = json.loads(line['errors'])

            for key in raw_headers:
                cell = cell_mapping[key.upper()] + str(line_counter)
                if survey:
                    try:
                        form_key, _ = get_mapping_from_csv_key(key)
                    except MappingNotFoundError:
                        form_key = key
                else:
                    form_key, _ = get_mapping_from_csv_key(key)

                if form_key in raw_data:

                    value = raw_data[form_key]

                    if value == "":
                        value = ''

                    ws1[cell] = value
                else:
                    ws1[cell] = ''

                if form_key in errors and form_key != 'city':
                    ws1[cell].fill = self.redFill
                    ws1[cell].comment = Comment(errors[form_key], 'Congressy')

            line_counter += 1

        wb.save(stream)

        return stream.getvalue()


class XLSLotExamplePersister(XLSPersister):

    def __init__(self, lot: Lot) -> None:
        if not isinstance(lot, Lot):
            raise TypeError('lot não é um objeto do tipo Lot')

        self.lot = lot
        self.bold = styles.Font(bold=True)
        super().__init__()

    def _get_header(self) -> list:

        headers = list()
        Header = namedtuple('Header',
                            ['title', 'required', 'possible_options', ])

        if hasattr(self.lot.event, 'formconfig'):
            form_config = self.lot.event.formconfig
        else:
            form_config = FormConfig()

        keys_mapping_list = get_keys_mapping_dict(form_config=form_config)
        for item in keys_mapping_list:
            headers.append(Header(
                title=item['mapping']['csv_keys'][0],
                required=item['required'],
                possible_options=list()
            ))

        survey = None
        if self.lot.event_survey:
            survey = self.lot.event_survey.survey

        if survey:
            questions = Question.objects.filter(
                survey=survey,
            ).order_by('order')

            for question in questions:
                cleaned_label = question.label.lower().strip()
                required = question.required

                options = question.options.all()
                options_list = list()
                for option in options:
                    options_list.append(option.name)

                headers.append(Header(
                    title=cleaned_label,
                    required=required,
                    possible_options=options_list
                ))

        return headers

    def make(self) -> bytes:

        headers_list = self._get_header()

        stream = BytesIO()

        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

        wb = Workbook()

        ws1 = wb.active

        # Some of the printable ASCII characters are invalid:  * : / \ ? [ ]
        name = self.lot.name \
            .replace('*', '') \
            .replace(':', '') \
            .replace('/', '-') \
            .replace("\\", '-') \
            .replace('?', '') \
            .replace('[', '') \
            .replace(']', '')

        ws1.title = 'Planilha de exemplo -  {}'.format(name)

        i = 1
        for header in headers_list:
            letter = self.colnum_string(i)
            cell = letter.upper() + str(1)
            ws1[cell] = header.title

            comment = None
            if header.required and len(header.possible_options) > 0:
                msg = 'Pergunta obrigatória\n'
                msg += 'Possiveis respostas\n\n'
                for opt in header.possible_options:
                    msg += opt + "\n"
                comment = Comment(msg, 'Congressy')
                comment.width = 300
                ws1[cell].font = self.bold
            elif header.required:
                comment = Comment('Pergunta obrigatória', 'Congressy')
                ws1[cell].font = self.bold
            elif len(header.possible_options) > 0:
                msg = 'Possiveis respostas\n\n'
                for opt in header.possible_options:
                    msg += opt + "\n"

                comment = Comment(msg, 'Congressy')
                comment.width = 300

            if comment:
                ws1[cell].comment = comment

            i += 1

        wb.save(stream)

        return stream.getvalue()
